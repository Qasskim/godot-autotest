# -*- coding: utf-8 -*-
"""테스트 결과 엑셀 기록. D:\\Automation\\RhythmGame_AutoTest\\common\\result_writer.py 방식 참고."""
import os
from datetime import datetime

import openpyxl
from openpyxl.drawing.image import Image as XLImage

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "reports", "test_results.xlsx")


def init_workbook():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
    else:
        wb = openpyxl.Workbook()

    sheet_name = _unique_sheet_name(wb, datetime.now().strftime("%d-%m-%Y"))
    ws = wb.create_sheet(title=sheet_name)
    wb.active = wb.sheetnames.index(sheet_name)

    # 새로 만든 워크북이면 비어있는 기본 시트("Sheet") 정리
    if "Sheet" in wb.sheetnames and wb["Sheet"] is not ws \
            and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 \
            and wb["Sheet"]["A1"].value is None:
        wb.remove(wb["Sheet"])

    ws.append(["test date/time", datetime.now().strftime("%x %X")])
    ws.append(["Test Number", "Test Case", "Result", "Note", "Screenshot"])
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 24
    _save(wb)
    return wb, ws


def _unique_sheet_name(wb, base_name: str) -> str:
    if base_name not in wb.sheetnames:
        return base_name
    n = 2
    while f"{base_name} ({n})" in wb.sheetnames:
        n += 1
    return f"{base_name} ({n})"


# (ws, row, path) 대기열 — 이미지는 세션 끝에 한 번만 삽입한다.
# openpyxl은 wb.save() 때마다 이미 삽입된 이미지까지 다시 직렬화하는데,
# PIL이 연 파일 핸들이 첫 save 이후 닫혀서 두 번째 save부터 "I/O operation on closed file"로 죽음.
_pending_images: list = []


def record(wb, ws, case_num: int, case_name: str, result: str, note: str = "",
           screenshot_path: str = None) -> None:
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=case_num)
    ws.cell(row=row, column=2, value=case_name)
    ws.cell(row=row, column=3, value=result)
    ws.cell(row=row, column=4, value=note)
    if screenshot_path and os.path.exists(screenshot_path):
        _pending_images.append((ws, row, screenshot_path))
        ws.row_dimensions[row].height = 70
    _save(wb)


def finalize(wb) -> None:
    """세션 종료 시 1회 호출 — 대기 중인 스크린샷을 전부 삽입하고 마지막으로 저장."""
    for ws, row, path in _pending_images:
        try:
            img = XLImage(path)
            img.width, img.height = 160, 90
            ws.add_image(img, f"E{row}")
        except Exception as e:
            print(f"\n[result_writer] 이미지 삽입 실패: {e}")
    _pending_images.clear()
    _save(wb)


def _save(wb) -> None:
    try:
        wb.save(EXCEL_PATH)
    except PermissionError:
        print(f"\n[result_writer] {EXCEL_PATH} 저장 실패 — 엑셀에서 파일을 열어놓은 상태면 닫고 다시 실행하세요.")
